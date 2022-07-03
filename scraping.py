from email.mime.multipart import MIMEMultipart
from selenium.webdriver.common.by import By
from PySimpleGUI import PySimpleGUI as sg
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from datetime import datetime
from email import encoders
from tqdm import tqdm
import smtplib
import json
import re
import os

from colors import Colors


class Scraping:
    
    def __init__(self, driver):
        self.driver = driver

        self.num_pages = By.XPATH, '/html/body/div[5]/div[2]/div[2]/div/div/nav/ul'
        self.btn_next = By.XPATH, '/html/body/div[5]/div[2]/div[2]/div/div/nav/ul/li[7]/a'
    
    def web_scraping(self):
        self.list_nome_phone = []
        self.list_preco_phone = []

        pages = self.driver.find_element(*self.num_pages).text
        pages = re.sub(r'\n', '', pages)
        if '«' in pages:
            pages = str(pages).strip('«').strip('»').strip()

        for page in tqdm (range(1, len(pages))):
            tqdm.write(f'Página: {page}')
            os.system('cls')
            for item in range(1, 13):
                nome_phone = self.driver.find_elements(By.XPATH, f'/html/body/div[5]/div[2]/div[1]/div[{item}]/div/h2/a')
                preco_phone = self.driver.find_elements(By.XPATH, f'/html/body/div[5]/div[2]/div[1]/div[{item}]/div/div[2]/ins')
                self.list_nome_phone.append(nome_phone[0].text)
                self.list_preco_phone.append(preco_phone[0].text)
            self.driver.find_element(*self.btn_next).click()
    
    def workbook(self, file='telefones importados.xlsx'):
        self.file = file
        if not os.path.isfile(self.file):
            print('não existe')
            self.wb = Workbook()
            self.wb.save(self.file)
        else:
            self.wb = load_workbook(self.file)
        self.ws = self.wb.active

    def inserir_dados(self, line, nome, preco):

        if  self.ws.cell(column=1, row=1).value == None:
            style_column_name = Font(name='Calibri',
                    size = 11,
                    bold = True)
            a1 = self.ws['A1']
            b1 = self.ws['B1']
            c1 = self.ws['C1']
            a1.font = style_column_name
            b1.font = style_column_name
            c1.font = style_column_name

            self.ws.cell(column=1, row=1).value = 'nome'
            self.ws.cell(column=2, row=1).value = 'preco'
            self.ws.cell(column=3, row=1).value = 'horário'

        self.ws.cell(column=1, row=line).value = nome
        self.ws.cell(column=2, row=line).value = preco
        self.ws.cell(column=3, row=line).value = datetime.now().strftime('%d/%m/%Y %H:%M')

    def window_email(self):

        sg.theme('Reddit')
        layout = [
            [sg.Text('gmail '), sg.Input(key='email', size=(35, 1))],
            [sg.Button('enviar', size=(8))]
        ]
        window = sg.Window('Gmail para envio do arquivo', layout, finalize=True)
        while True:
            event, self.value = window.read()
            if event == 'enviar':
                break
            elif event == sg.WIN_CLOSED:
                break
        window.close()
        
    def send_email(self):

        with open('server.json', 'r') as f:
            server = json.load(f)
        login = server['login']
        password = server['password']

        server = smtplib.SMTP(server['host'], server['port'])
        server.ehlo()
        server.starttls()
        server.login(login, password)

        email_msg = MIMEMultipart()
        email_msg['From'] = login
        email_msg['To'] = self.value['email']
        email_msg['Subject'] = 'Planilha Phones e-commerce'

        email_msg.attach(MIMEText('Olá. <br> Segue anexo arquivo <b>.xlsx</b> com os nomes e preços correspondentes \
            aos celulares do site: https://telefonesimportados.netlify.app/', 'html'))

        attachment = open(self.file, 'rb')
        att = MIMEBase('application', 'octect-stream')
        att.set_payload(attachment.read())
        encoders.encode_base64(att)

        att.add_header('Content-Disposition', f'attachment; filename={self.file}')
        attachment.close()

        email_msg.attach(att)
        try:
            server.sendmail(email_msg['From'], email_msg['To'], email_msg.as_string())
            server.quit()
            print('\nEmail successfully sent.')
        except Exception as e:
            print(f'{Colors.red}ERROR: {Colors.reset}{e}')
 